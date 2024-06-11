import torch.optim as optim
import torch.nn.functional as F
import openpyxl

from utils.utility import *
from utils.metric import *
from utils.approximate import *

### Trainer for 'SVM'
class SVM_Trainer:
    def __init__(self, args, device, network, train_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.test_loader = test_loader
    
    def train(self):
        for adjacency, feature, label in self.train_loader:
            self.network.fit(feature, label)
    
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        for adjacency, feature, label in self.test_loader:
            output = self.network.predict(feature)
            output = torch.FloatTensor(encode_onehot(output))

            loss_test = torch.tensor([0])
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)

            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"One vs One - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


class MLP_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
        
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                # import pdb
                # pdb.set_trace()
                output = self.network.forward(feature) # Shape: (# of samples, # of labels)
                
                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                # import pdb
                # pdb.set_trace()
                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            output = self.network.forward(feature) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'GCN', 'GAT', 'GDC'
class GNN_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)

    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'GraphHeat'
class GraphHeat_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_sz):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_sz = adj_sz
        self.optimizer = optimizer

        if args.use_t_local == 1: 
            self.t = torch.empty(adj_sz).fill_(2.)
        else:
            self.t = torch.tensor([2.])
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)
            
            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None
    

### Trainer for 'Ours'
class Exact_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_size):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_size = adj_size
        self.optimizer = optimizer

        if args.use_t_local == 1: # Local scale
            self.t = torch.empty(adj_size).fill_(2.)
        else: # Global scale
            self.t = torch.tensor([2.]) 
        
        self.t_lr = self.args.t_lr
        self.t_loss_threshold = self.args.t_loss_threshold
        self.t_lambda = self.args.t_lambda
        self.t_threshold = self.args.t_threshold
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    ### Scale regularization of loss function
    def t_loss(self):
        t_one = torch.abs(self.t)
        t_zero = torch.zeros_like(self.t)

        t_l = torch.where(self.t < self.t_loss_threshold, t_one, t_zero)

        return self.t_lambda * torch.sum(t_l)

    def t_deriv(self):
        t_one = self.t_lambda * torch.ones_like(self.t)
        t_zero = torch.zeros_like(self.t)

        t_de = torch.where(self.t < self.t_loss_threshold, -t_one, t_zero)

        return t_de

    def scale_update(self, output, feature, label, heat_kernel, heat_kernel_grad):
        y_oh = torch.zeros_like(output) # (# sample, # label)
        y_oh.scatter_(1, label.reshape(-1, 1), 1)
        dl_ds = (torch.exp(output) - y_oh) / output.shape[0]
        
        ds_dro0 = torch.mul(dl_ds, self.network.linrdp2) @ self.network.linear2.weight
        ds_dro1 = torch.mul(ds_dro0, self.network.linrdp)
        
        #ds_dro1 = torch.mul(dl_ds @ self.network.linear2.weight,  self.network.linrdp)
        dl_dro = torch.matmul(ds_dro1, self.network.linear.weight).reshape(-1, heat_kernel.shape[-2], self.args.hidden_units)
        
        if self.args.layer_num == 1:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp) @ self.network.gcn.gc1.weight.T
            dl_dt = torch.mul((dl_dc @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
        elif self.args.layer_num == 2:
            dl_dl2 = torch.mul(dl_dro, self.network.gcn.rdp2) @ self.network.gcn.gc2.weight.T

            dl_first = torch.mul((dl_dl2 @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
            backward = torch.matmul(self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2))

            dl_second_tmp = torch.mul(dl_dl2, self.network.gcn.rdp)
            dl_second = torch.matmul(torch.mul(dl_second_tmp @ backward, heat_kernel_grad), heat_kernel.swapaxes(1, 2))

            dl_dt = dl_first + dl_second
        elif self.args.layer_num == 3:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp3) @ self.network.gcn.gc3.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third
        elif self.args.layer_num == 4:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp4) @ self.network.gcn.gc4.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f3.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_fourth = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp)
                            @ self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2)), heat_kernel_grad),
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third + dl_dc_fourth

        if self.args.use_t_local == 1:
            dl_dt = torch.sum(dl_dt, dim=(0, 2))
        else:
            dl_dt = torch.tensor([torch.sum(dl_dt, dim=(0, 1, 2))]).to(self.device)
            
        dl_dt += self.t_deriv() # Add regularizer on t
        now_lr = self.t_lr * dl_dt

        now_lr[now_lr > self.t_threshold] = self.t_threshold
        now_lr[now_lr < -self.t_threshold] = -self.t_threshold

        self.t = self.t - now_lr # Update t

        if self.args.use_t_local == 1:
            print(f't:{self.t[0].item()}', end=' ')
        else:
            print(f't:{self.t.item():.4f}', end=' ')

    ### Train
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="epoch"+str(0))
        for y in range(2, self.t.shape[0] + 2):
            ws.cell(row=1, column=y, value=2)
        i = 2
        
        for epoch in range(1, self.args.epochs + 1):
            self.network.train()
            
            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero

                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label) + self.t_loss()
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                with torch.no_grad():
                    self.scale_update(output, feature, label, heat_kernel, heat_kernel_grad)
                
                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())
                
                s = self.t.detach().cpu().numpy()
                ws.cell(row=i, column=1, value="epoch"+str(epoch))
                for y in range(2, self.t.shape[0] + 2):
                    ws.cell(row=i, column=y, value=s[y-2])
                i += 1
            
            #wb.save("/home/user/ASAP_new/GC/our_s.xlsx")
            self.network.eval()
        

    ### Test
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.eval()

        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)

            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label) + self.t_loss()
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            for i in range(len(self.t)):
                ts.append(self.t[i].item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)
    

### Trainer for approximation version of 'Ours'
class LSAP_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_size):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_size = adj_size # 160
        self.optimizer = optimizer

        if args.use_t_local == 1: # Local scale 
            self.t = torch.empty(adj_size).fill_(2.) # (160)
        else: # Global scale 
            self.t = torch.tensor([2.]) # (1)
        
        self.t_lr = self.args.t_lr # 1
        self.t_loss_threshold = self.args.t_loss_threshold # 0.01
        self.t_lambda = self.args.t_lambda # 1
        self.t_threshold = self.args.t_threshold # 0.1
        
        self.m_che = self.args.m_chebyshev
        self.m_her = self.args.m_hermite 
        self.m_lag = self.args.m_laguerre 
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    ### Scale regularization of loss function
    def t_loss(self):
        t_one = torch.abs(self.t)
        t_zero = torch.zeros_like(self.t)

        t_loss_tmp = torch.where(self.t < self.t_loss_threshold, t_one, t_zero)
        t_loss_final = self.t_lambda * torch.sum(t_loss_tmp) # λ|s|

        return t_loss_final

    def t_deriv(self):
        t_one = self.t_lambda * torch.ones_like(self.t)
        t_zero = torch.zeros_like(self.t)

        t_de = torch.where(self.t < self.t_loss_threshold, -t_one, t_zero)

        return t_de

    def scale_update(self, output, feature, label, heat_kernel, heat_kernel_grad):
        y_oh = torch.zeros_like(output) # (# sample, # label)
        y_oh.scatter_(1, label.reshape(-1, 1), 1) # (# sample, # label)

        dl_ds = (torch.exp(output) - y_oh) / output.shape[0] # (# sample, # label)

        ds_dro0 = torch.mul(dl_ds, self.network.linrdp2) @ self.network.linear2.weight
        ds_dro1 = torch.mul(ds_dro0, self.network.linrdp)
        
        #ds_dro1 = torch.mul(dl_ds @ self.network.linear2.weight,  self.network.linrdp)
        dl_dro = torch.matmul(ds_dro1, self.network.linear.weight).reshape(-1, heat_kernel.shape[-2], self.args.hidden_units)
        
        if self.args.layer_num == 1:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp) @ self.network.gcn.gc1.weight.T
            dl_dt = torch.mul((dl_dc @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
        elif self.args.layer_num == 2:
            dl_dl2 = torch.mul(dl_dro, self.network.gcn.rdp2) @ self.network.gcn.gc2.weight.T

            dl_first = torch.mul((dl_dl2 @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
            backward = torch.matmul(self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2))

            dl_second_tmp = torch.mul(dl_dl2, self.network.gcn.rdp)
            dl_second = torch.matmul(torch.mul(dl_second_tmp @ backward, heat_kernel_grad), heat_kernel.swapaxes(1, 2))

            dl_dt = dl_first + dl_second
        elif self.args.layer_num == 3:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp3) @ self.network.gcn.gc3.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third
        elif self.args.layer_num == 4:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp4) @ self.network.gcn.gc4.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f3.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_fourth = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp)
                            @ self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2)), heat_kernel_grad),
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third + dl_dc_fourth

        if self.args.use_t_local == 1:
            dl_dt = torch.sum(dl_dt, dim=(0, 2)) # 160
        else:
            dl_dt = torch.tensor([torch.sum(dl_dt, dim=(0, 1, 2))]).to(self.device) # 1

        dl_dt += self.t_deriv() # 160 / 1
        now_lr = self.t_lr * dl_dt # 160 / 1
        
        now_lr[now_lr > self.t_threshold] = self.t_threshold
        now_lr[now_lr < -self.t_threshold] = -self.t_threshold
        
        self.t = self.t - now_lr # Update t 

        if self.args.use_t_local == 1:
            print(f't:{self.t[0].item()}', end=' ')
        else:
            print(f't:{self.t.item():.4f}', end=' ')

    ### Train
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="epoch"+str(0))
        for y in range(2, self.t.shape[0] + 2):
            ws.cell(row=1, column=y, value=2)
        i = 2

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()
            
            for adjacency, feature, label, eigenvalue, eigenvector, laplacian, P_n in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                if self.args.polynomial == 0: # Chebyshev
                    b = eigenvalue[:,eigenvalue.shape[1]-1]
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_chebyshev(P_n, laplacian, self.m_che, self.t, b, self.device)
                elif self.args.polynomial == 1: # Hermite
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_hermite(P_n, laplacian, self.m_her, self.t, self.device)
                elif self.args.polynomial == 2: # Laguerre
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_laguerre(P_n, laplacian, self.m_lag, self.t, self.device)

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label) + self.t_loss()
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                with torch.no_grad():
                    self.scale_update(output, feature, label, heat_kernel, heat_kernel_grad)
                
                self.optimizer.step() # Updates the parameters
                
                #wandb.log({"loss_train": loss_train.item(),
                #           "accuracy_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())
                
                s = self.t.detach().cpu().numpy()
                ws.cell(row=i, column=1, value="epoch"+str(epoch))
                for y in range(2, self.t.shape[0] + 2):
                    ws.cell(row=i, column=y, value=s[y-2])
                i += 1
            
            #wb.save("/home/user/ASAP_new/GC/our_s_"+ str(self.args.polynomial) +".xlsx")
            
            self.network.eval()
            
    ### Test
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.eval()

        for adjacency, feature, label, eigenvalue, eigenvector, laplacian, P_n in self.test_loader:
            if self.args.polynomial == 0: # Chebyshev
                b = eigenvalue[:,eigenvalue.shape[1]-1]
                heat_kernel, _ = compute_heat_kernel_chebyshev(P_n, laplacian, self.m_che, self.t, b, self.device)
            elif self.args.polynomial == 1: # Hermite
                heat_kernel, _ = compute_heat_kernel_hermite(P_n, laplacian, self.m_her, self.t, self.device)
            elif self.args.polynomial == 2: # Laguerre
                heat_kernel, _ = compute_heat_kernel_laguerre(P_n, laplacian, self.m_lag, self.t, self.device)

            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label) + self.t_loss()
            accuracy_test = compute_accuracy(output, label)
            
            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            for i in range(len(self.t)):
                ts.append(self.t[i].item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)